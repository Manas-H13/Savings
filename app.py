from flask import Flask, render_template, request, jsonify, send_file
import os
import random
from datetime import datetime
import openpyxl
from openpyxl.styles import Font

app = Flask(__name__)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

EXCEL_FILE = 'expenses.xlsx'
CATEGORIES = ['Travel', 'Petrol', 'Food', 'Groceries', 'Utilities', 'Shopping', 'Other']
COLUMNS = ['Date'] + CATEGORIES + ['Total']

def read_from_excel():
    if not os.path.exists(EXCEL_FILE): return {}
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    
    data = {}
    header_found = False
    col_map = {}
    current_month_str = None
    
    for row_cells in ws.iter_rows(values_only=True):
        if not row_cells or row_cells[0] is None or str(row_cells[0]).strip() == "": continue
        
        cell_val = str(row_cells[0]).strip()
        
        # Check if it's a month header (e.g., April 2026)
        try:
            d_month = datetime.strptime(cell_val, "%B %Y")
            current_month_str = d_month.strftime("%Y-%m")
            continue
        except ValueError:
            pass
            
        # Is it a Column Header?
        if cell_val in ('Date', 'Sl No.'):
            col_map = {col: i for i, col in enumerate(row_cells) if col}
            continue
            
        # Parse day / date
        date_val = None
        try:
            if "-" in cell_val and len(cell_val) >= 10:
                d = datetime.strptime(cell_val[:10], "%Y-%m-%d")
                date_val = d.strftime("%Y-%m-%d")
                current_month_str = d.strftime("%Y-%m")
            elif current_month_str:
                day_num = int(float(cell_val))
                date_val = f"{current_month_str}-{day_num:02d}"
                datetime.strptime(date_val, "%Y-%m-%d") # Validate
        except ValueError:
            continue
            
        if not date_val: continue
        
        if date_val not in data:
            data[date_val] = {cat: 0.0 for cat in CATEGORIES}
            
        for cat in CATEGORIES:
            idx = col_map.get(cat)
            if idx is not None and idx < len(row_cells):
                val = row_cells[idx]
                if val:
                    try: 
                        data[date_val][cat] += float(val)
                    except: pass
                    
    return data

def write_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    sorted_dates = sorted(data.keys())
    current_month = None
    
    if not sorted_dates:
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    for date_val in sorted_dates:
        d = datetime.strptime(date_val, "%Y-%m-%d")
        month_label = d.strftime("%B %Y")
        
        if month_label != current_month:
            if current_month is not None:
                ws.append([]) # space
            ws.append([month_label])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14, color="0000FF")
            
            ws.append(COLUMNS)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                
            current_month = month_label
            
        # Use Day number for Sl No.
        row = [int(d.strftime("%d"))]
        total = 0.0
        for cat in CATEGORIES:
            amt = data[date_val].get(cat, 0.0)
            row.append(amt if amt > 0 else "")
            total += amt
        row.append(total)
        ws.append(row)
        
    wb.save(EXCEL_FILE)

if not os.path.exists(EXCEL_FILE):
    write_to_excel({})

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download-excel')
def download_excel():
    return send_file(EXCEL_FILE, as_attachment=True)

@app.route('/api/expenses', methods=['GET', 'POST'])
def handle_expenses():
    if request.method == 'POST':
        req_data = request.json
        date = req_data.get('date', datetime.today().strftime('%Y-%m-%d'))
        amount = float(req_data.get('amount', 0))
        category = req_data.get('category', 'Other')

        data = read_from_excel()
        if date not in data:
            data[date] = {cat: 0.0 for cat in CATEGORIES}
        
        if category in data[date]:
            data[date][category] += amount
            
        write_to_excel(data)
        return jsonify({'status': 'success'})

    else:
        # GET expenses
        data = read_from_excel()
        results = []
        for d_key, cats in data.items():
            entry = {'Date': d_key}
            total = 0
            for c, a in cats.items():
                entry[c] = a
                total += a
            entry['Total'] = total
            results.append(entry)
            
        return jsonify(results)

@app.route('/api/chat', methods=['POST'])
def chat():
    user_msg = request.json.get('message', '').lower()
    data = read_from_excel()
    
    cat_totals = {cat: 0.0 for cat in CATEGORIES}
    total_spent = 0.0
    for cats in data.values():
        for c, a in cats.items():
            cat_totals[c] += a
            total_spent += a
            
    top_cat = max(cat_totals, key=cat_totals.get) if total_spent > 0 else "None"
    top_amt = cat_totals.get(top_cat, 0)
    
    reply = ""
    
    # 1. Full Details / Expenses
    if any(word in user_msg for word in ["expense", "detail", "summary", "breakdown", "full", "report"]):
        if total_spent == 0:
            reply = "You currently have ₹0 logged! Add some transactions to get a full report."
        else:
            breakdown_list = [f"{c}: ₹{a:.2f}" for c, a in cat_totals.items() if a > 0]
            breakdown_text = ", ".join(breakdown_list)
            reply = f"Here is your full financial breakdown: {breakdown_text}. Your absolute total spent globally is ₹{total_spent:.2f}. "
            if top_amt > 0:
                reply += f"Remember, {top_cat} makes up the bulk of this at ₹{top_amt:.2f}."
                
    # 2. Estimation
    elif "estimate" in user_msg or "amount" in user_msg or "if saved" in user_msg:
        potential = total_spent * 0.20
        if total_spent > 0:
            reply = f"Based on your current logged expenses, if you cut back 20%, I estimate you can save around ₹{potential:.2f}! To achieve this, automate a ₹{potential:.2f} transfer to your secure savings instantly after earning."
        else:
            reply = "I estimate ₹0 right now! Log some data so I can give you an accurate estimation."
            
    # 3. Achieve Goals
    elif "achieve" in user_msg or "goal" in user_msg:
        reply = "To achieve your numbers, follow the 50/30/20 rule: 50% for Needs, 30% for Wants, and 20% into Savings. Review your transactions tab daily to stay strictly within limits!"
        
    # 4. Savings Logic
    elif "save" in user_msg or "how" in user_msg:
        if total_spent == 0:
            reply = "You haven't spent anything yet! Keep tracking to unlock deep insights."
        else:
            reply = f"The fastest way for you to save right now is by putting a hard cap on {top_cat}. You've spent ₹{top_amt:.2f} there! Think carefully before buying more in {top_cat}."
            
    # 5. General Conversation
    elif any(word in user_msg for word in ["hi", "hello", "hey"]):
        reply = "Hello! I am Hai, your personal AI savings advisor. How can I help you manage your finances today?"
    elif any(word in user_msg for word in ["who", "what are you", "can you"]):
        reply = "I am Hai! I analyze your native Excel data to provide tailored savings strategies, estimate your targets accurately, and generate full analytical breakdowns of your spending line-by-line."
    elif "thank" in user_msg:
        reply = "You're very welcome! Let's hit those financial goals together!"
    else:
        reply = "I'm Hai! Try asking me for 'full expense details', 'How can I save?', 'estimated savings', or just say Hello!"
        
    return jsonify({"reply": reply})

if __name__ == '__main__':
    app.run(debug=True, port=8080)
